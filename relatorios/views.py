from datetime import date, timedelta
from decimal import Decimal

from dateutil.relativedelta import relativedelta
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, F, Q, Sum
from django.db.models.functions import Coalesce, ExtractHour, ExtractWeekDay, TruncDay, TruncMonth, TruncWeek, TruncYear
from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone

from catalogo.models import Produto
from financeiro.models import CaixaSessao, ContaPagar, ContaReceber, MovimentoCaixa
from vendas.models import Cliente, Devolucao, ItemVenda, PerfilVendedor, Venda

from .models import MetaMensal, PreferenciaDashboard, get_config_estoque
from .widgets_registry import montar_lista_widgets

MESES_PT = [
    "jan", "fev", "mar", "abr", "mai", "jun",
    "jul", "ago", "set", "out", "nov", "dez",
]


# Exportação em Excel reaproveita os mesmos dados do dashboard.
@user_passes_test(lambda u: u.is_staff, login_url="/login/")
def exportar_excel(request):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font

    contexto = montar_contexto_dashboard()
    wb = openpyxl.Workbook()

    # --- Resumo -------------------------------------------------------------
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Belletti Cards Universe — Dashboard de vendas"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    linhas_resumo = [
        ("Total de vendas (12 meses)", contexto["total_valor"]),
        ("Ticket médio", contexto["ticket_medio"]),
        ("Pedidos fechados", contexto["total_pedidos"]),
        ("Itens vendidos", contexto["produtos_vendidos"]),
        ("Produtos em catálogo", contexto["total_produtos_catalogo"]),
        ("Saldo atual em caixa", contexto["saldo_atual"]),
        ("A receber (pendente)", contexto["total_a_receber"]),
    ]
    for label, valor in linhas_resumo:
        ws.append([label, valor])

    # --- Vendas por mês ------------------------------------------------------
    ws2 = wb.create_sheet("Vendas por mês")
    ws2.append(["Mês", "Valor (R$)"])
    for label, valor in zip(contexto["labels_mes"], contexto["valores_mes"]):
        ws2.append([label, valor])

    # --- DRE ------------------------------------------------------------------
    ws3 = wb.create_sheet("DRE (mês atual)")
    dre = contexto["dre"]
    ws3.append(["Linha", "Valor (R$)"])
    for label, chave in [
        ("Receita bruta", "receita_bruta"), ("Devoluções", "valor_devolucoes"),
        ("Receita líquida", "receita_liquida"), ("CMV", "cmv"),
        ("Margem bruta", "margem_bruta"), ("Despesas operacionais", "despesas_operacionais"),
        ("Lucro líquido", "lucro_liquido"),
    ]:
        ws3.append([label, dre[chave]])

    # --- Produtos mais vendidos / Curva ABC -----------------------------------
    ws4 = wb.create_sheet("Produtos (Curva ABC)")
    ws4.append(["Produto", "Valor (R$)", "% acumulado", "Classe"])
    for p in contexto["curva_abc"]:
        ws4.append([p["nome"], p["valor"], float(p["percentual_acumulado"]), p["classe"]])

    # --- Comissões --------------------------------------------------------------
    ws5 = wb.create_sheet("Comissões")
    ws5.append(["Vendedor", "Vendido (R$)", "% comissão", "Comissão (R$)"])
    for c in contexto["comissoes"]:
        ws5.append([c["vendedor"], c["total_vendido"], float(c["percentual"]), c["comissao"]])

    # --- Próximos vencimentos -----------------------------------------------
    ws6 = wb.create_sheet("Próximos vencimentos")
    ws6.append(["Tipo", "Descrição", "Vencimento", "Valor (R$)"])
    for v in contexto["proximos_vencimentos"]:
        ws6.append([v["tipo"], v["descricao"], v["vencimento"].strftime("%d/%m/%Y"), v["valor"]])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nome_arquivo = f"dashboard_belletti_{timezone.localdate():%Y%m%d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)
    return response


# Dashboard mostra faturamento, margem e caixa — informação sensível
# que só quem tem acesso staff (dono/gerência) deve ver. Vendedor fica
# só com o PDV.
@user_passes_test(lambda u: u.is_staff, login_url="/login/")
def dashboard(request):
    contexto = montar_contexto_dashboard()
    preferencia = PreferenciaDashboard.objects.filter(usuario=request.user).first()
    contexto["widgets"] = montar_lista_widgets(preferencia)
    return render(request, "relatorios/dashboard.html", contexto)


@user_passes_test(lambda u: u.is_staff, login_url="/login/")
def salvar_preferencia_dashboard(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método não permitido."}, status=405)

    import json
    try:
        dados = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "erro": "Dados inválidos."}, status=400)

    from .widgets_registry import WIDGETS_POR_ID
    ordem = [wid for wid in dados.get("ordem", []) if wid in WIDGETS_POR_ID]
    ocultos = [wid for wid in dados.get("ocultos", []) if wid in WIDGETS_POR_ID]

    PreferenciaDashboard.objects.update_or_create(
        usuario=request.user,
        defaults={"ordem_widgets": ordem, "widgets_ocultos": ocultos},
    )
    return JsonResponse({"ok": True})


def montar_contexto_dashboard():
    """
    Monta todos os dados do dashboard. Extraído numa função à parte
    (em vez de ficar só dentro da view) pra poder ser reaproveitado
    também na exportação pra Excel, sem duplicar a lógica.
    """
    hoje = timezone.localdate()
    inicio_periodo = (hoje.replace(day=1) - relativedelta(months=11))

    vendas_fechadas = Venda.objects.filter(status="fechada")
    vendas_periodo = vendas_fechadas.filter(fechada_em__date__gte=inicio_periodo)

    # --- Cards de visão geral -------------------------------------------------
    agregados = vendas_periodo.aggregate(
        total_valor=Sum(F("itens__quantidade") * F("itens__preco_unitario")),
        total_pedidos=Count("id", distinct=True),
    )
    total_valor = agregados["total_valor"] or 0
    total_pedidos = agregados["total_pedidos"] or 0
    ticket_medio = (total_valor / total_pedidos) if total_pedidos else 0

    produtos_vendidos = (
        ItemVenda.objects.filter(venda__in=vendas_periodo)
        .aggregate(total=Sum("quantidade"))["total"] or 0
    )

    contas_pendentes = ContaReceber.objects.filter(status="pendente")
    total_a_receber = contas_pendentes.aggregate(total=Sum("valor"))["total"] or 0
    qtd_a_receber = contas_pendentes.count()

    # --- Gráfico: vendas por mês (últimos 12 meses) ---------------------------
    por_mes = (
        vendas_periodo
        .annotate(mes=TruncMonth("fechada_em"))
        .values("mes")
        .annotate(
            valor=Sum(F("itens__quantidade") * F("itens__preco_unitario")),
            pedidos=Count("id", distinct=True),
        )
        .order_by("mes")
    )
    mapa_mes = {item["mes"].date(): item for item in por_mes if item["mes"]}

    labels_mes = []
    valores_mes = []
    cursor = inicio_periodo
    for _ in range(12):
        item = mapa_mes.get(cursor)
        labels_mes.append(f"{MESES_PT[cursor.month - 1]}/{str(cursor.year)[2:]}")
        valores_mes.append(float(item["valor"]) if item else 0)
        cursor = cursor + relativedelta(months=1)

    # --- Top produtos mais vendidos --------------------------------------------
    top_produtos = (
        ItemVenda.objects.filter(venda__in=vendas_periodo)
        .values("produto__nome", "produto__sku")
        .annotate(
            quantidade_total=Sum("quantidade"),
            valor_total=Sum(F("quantidade") * F("preco_unitario")),
        )
        .order_by("-quantidade_total")[:5]
    )

    # --- Produtos menos vendidos (inclui os que nunca venderam nenhuma vez) -----
    produtos_menos_vendidos = (
        Produto.objects.filter(ativo=True)
        .annotate(
            quantidade_total_venda=Coalesce(
                Sum("itemvenda__quantidade", filter=Q(itemvenda__venda__in=vendas_periodo)), 0
            ),
            valor_total_venda=Coalesce(
                Sum(
                    F("itemvenda__quantidade") * F("itemvenda__preco_unitario"),
                    filter=Q(itemvenda__venda__in=vendas_periodo),
                ), Decimal("0"),
            ),
        )
        .order_by("quantidade_total_venda", "nome")[:8]
    )

    # --- Horários com mais vendas ----------------------------------------------
    tz_atual = timezone.get_current_timezone()
    por_hora = (
        vendas_periodo
        .annotate(hora=ExtractHour("fechada_em", tzinfo=tz_atual))
        .values("hora")
        .annotate(total=Count("id"))
        .order_by("-total")[:8]
    )
    labels_hora = [f"{item['hora']:02d}:00" for item in por_hora]
    valores_hora = [item["total"] for item in por_hora]

    # --- Fluxo de caixa projetado (próximos 30 dias) ---------------------------
    entradas_realizadas = MovimentoCaixa.objects.filter(tipo="entrada").aggregate(
        total=Sum("valor"))["total"] or Decimal("0")
    saidas_realizadas = MovimentoCaixa.objects.filter(tipo="saida").aggregate(
        total=Sum("valor"))["total"] or Decimal("0")
    saldo_atual = entradas_realizadas - saidas_realizadas

    fim_projecao = hoje + timedelta(days=30)
    receber_por_dia = {
        item["vencimento"]: item["total"]
        for item in ContaReceber.objects.filter(
            status="pendente", vencimento__gte=hoje, vencimento__lte=fim_projecao
        ).values("vencimento").annotate(total=Sum("valor"))
    }
    pagar_por_dia = {
        item["vencimento"]: item["total"]
        for item in ContaPagar.objects.filter(
            status="pendente", vencimento__gte=hoje, vencimento__lte=fim_projecao
        ).values("vencimento").annotate(total=Sum("valor"))
    }

    labels_fluxo = []
    valores_fluxo = []
    saldo_corrente = saldo_atual
    for i in range(31):
        dia = hoje + timedelta(days=i)
        saldo_corrente += receber_por_dia.get(dia, Decimal("0"))
        saldo_corrente -= pagar_por_dia.get(dia, Decimal("0"))
        labels_fluxo.append(dia.strftime("%d/%m"))
        valores_fluxo.append(float(saldo_corrente))

    proximos_vencimentos = sorted(
        [
            {"tipo": "A receber", "descricao": c.descricao, "valor": c.valor, "vencimento": c.vencimento}
            for c in ContaReceber.objects.filter(status="pendente", vencimento__lte=fim_projecao)
        ] + [
            {"tipo": "A pagar", "descricao": c.descricao, "valor": c.valor, "vencimento": c.vencimento}
            for c in ContaPagar.objects.filter(status="pendente", vencimento__lte=fim_projecao)
        ],
        key=lambda x: x["vencimento"],
    )[:10]

    # --- DRE simplificado do mês atual ------------------------------------------
    inicio_mes = hoje.replace(day=1)
    vendas_mes = Venda.objects.filter(status="fechada", fechada_em__date__gte=inicio_mes)

    receita_bruta = vendas_mes.aggregate(
        total=Sum(F("itens__quantidade") * F("itens__preco_unitario"))
    )["total"] or Decimal("0")
    total_descontos = vendas_mes.aggregate(total=Sum("desconto"))["total"] or Decimal("0")
    receita_bruta -= total_descontos

    devolucoes_mes = Devolucao.objects.filter(
        processada=True, processada_em__date__gte=inicio_mes, venda__in=vendas_fechadas
    )
    valor_devolucoes = sum((d.total for d in devolucoes_mes), Decimal("0"))

    receita_liquida = receita_bruta - valor_devolucoes

    cmv = ItemVenda.objects.filter(venda__in=vendas_mes).aggregate(
        total=Sum(F("quantidade") * F("produto__preco_custo"))
    )["total"] or Decimal("0")

    margem_bruta = receita_liquida - cmv

    despesas_operacionais = ContaPagar.objects.filter(
        vencimento__gte=inicio_mes, vencimento__lte=hoje
    ).exclude(categoria__nome="Compra de estoque").aggregate(
        total=Sum("valor"))["total"] or Decimal("0")

    lucro_liquido = margem_bruta - despesas_operacionais

    # --- Indicadores financeiros --------------------------------------------------
    def _pct(numerador, denominador):
        if not denominador:
            return None
        return round((numerador / denominador) * 100, 1)

    indicadores = {
        "margem_bruta_pct": _pct(margem_bruta, receita_liquida),
        "margem_liquida_pct": _pct(lucro_liquido, receita_liquida),
        "cmv_pct": _pct(cmv, receita_liquida),
        "despesas_pct": _pct(despesas_operacionais, receita_liquida),
        "ponto_equilibrio": None,
    }
    if margem_bruta > 0 and receita_liquida > 0:
        margem_pct_decimal = margem_bruta / receita_liquida
        indicadores["ponto_equilibrio"] = (despesas_operacionais / margem_pct_decimal).quantize(Decimal("0.01"))

    # --- Produtos diferentes em catálogo ----------------------------------------
    total_produtos_catalogo = Produto.objects.filter(ativo=True).count()

    # --- Meta do mês -------------------------------------------------------------
    meta_mensal = MetaMensal.objects.filter(mes=inicio_mes).first()
    meta_contexto = None
    if meta_mensal and meta_mensal.valor > 0:
        progresso_percentual = min(
            (receita_bruta / meta_mensal.valor) * 100, 100
        )
        falta_valor = max(meta_mensal.valor - receita_bruta, Decimal("0"))
        falta_percentual = max(100 - progresso_percentual, 0)
        meta_contexto = {
            "valor_meta": meta_mensal.valor,
            "progresso_percentual": progresso_percentual,
            "falta_valor": falta_valor,
            "falta_percentual": falta_percentual,
            "atingida": receita_bruta >= meta_mensal.valor,
        }

    # --- Produtos que precisam de reestoque --------------------------------------
    produtos_reestoque = (
        Produto.objects.filter(ativo=True, estoque_atual__lte=F("estoque_minimo"))
        .order_by("estoque_atual")[:8]
    )

    # --- Produtos parados (sem venda há muito tempo) ----------------------------
    config_estoque = get_config_estoque()
    data_limite_parado = hoje - timedelta(days=config_estoque.dias_produto_parado)

    produtos_vendidos_recentemente = (
        ItemVenda.objects.filter(
            venda__status="fechada", venda__fechada_em__date__gte=data_limite_parado
        ).values_list("produto_id", flat=True).distinct()
    )
    produtos_parados_qs = (
        Produto.objects.filter(ativo=True, estoque_atual__gt=0)
        .exclude(id__in=produtos_vendidos_recentemente)
        .exclude(criado_em__date__gte=data_limite_parado)  # dá um tempo pra produto novo
        .order_by("-estoque_atual")[:8]
    )
    produtos_parados = []
    for p in produtos_parados_qs:
        ultima_venda = (
            ItemVenda.objects.filter(produto=p, venda__status="fechada")
            .order_by("-venda__fechada_em").first()
        )
        produtos_parados.append({
            "produto": p,
            "ultima_venda": ultima_venda.venda.fechada_em.date() if ultima_venda else None,
        })

    # --- Comissão por vendedor (mês atual) ---------------------------------------
    comissoes = []
    vendas_por_vendedor = (
        vendas_mes.exclude(vendedor__isnull=True)
        .values("vendedor__username", "vendedor_id")
        .annotate(total_vendido=Sum(F("itens__quantidade") * F("itens__preco_unitario")))
        .order_by("-total_vendido")
    )
    percentuais = {
        p.usuario_id: p.percentual_comissao
        for p in PerfilVendedor.objects.all()
    }
    for linha in vendas_por_vendedor:
        percentual = percentuais.get(linha["vendedor_id"], Decimal("0"))
        comissoes.append({
            "vendedor": linha["vendedor__username"],
            "total_vendido": linha["total_vendido"] or Decimal("0"),
            "percentual": percentual,
            "comissao": (linha["total_vendido"] or Decimal("0")) * percentual / 100,
        })

    # --- Curva ABC de produtos (últimos 12 meses) --------------------------------
    produtos_valor = (
        ItemVenda.objects.filter(venda__in=vendas_periodo)
        .values("produto__nome")
        .annotate(valor=Sum(F("quantidade") * F("preco_unitario")))
        .order_by("-valor")
    )
    valor_total_periodo = sum((p["valor"] for p in produtos_valor), Decimal("0"))
    curva_abc = []
    acumulado = Decimal("0")
    for p in produtos_valor:
        acumulado += p["valor"]
        percentual_acumulado = (acumulado / valor_total_periodo * 100) if valor_total_periodo else Decimal("0")
        if percentual_acumulado <= 80:
            classe = "A"
        elif percentual_acumulado <= 95:
            classe = "B"
        else:
            classe = "C"
        curva_abc.append({
            "nome": p["produto__nome"], "valor": p["valor"],
            "percentual_acumulado": percentual_acumulado, "classe": classe,
        })

    # --- Comparação mensal e anual (mesmo nº de dias decorridos, pra ser justo) --
    def _receita_bruta_periodo(data_inicio, data_fim):
        vendas = Venda.objects.filter(
            status="fechada", fechada_em__date__gte=data_inicio, fechada_em__date__lte=data_fim
        )
        bruta = vendas.aggregate(
            total=Sum(F("itens__quantidade") * F("itens__preco_unitario"))
        )["total"] or Decimal("0")
        desconto = vendas.aggregate(total=Sum("desconto"))["total"] or Decimal("0")
        return bruta - desconto

    def _variacao_percentual(atual, anterior):
        if anterior > 0:
            return ((atual - anterior) / anterior) * 100
        return Decimal("100") if atual > 0 else Decimal("0")

    dias_decorridos_mes = (hoje - inicio_mes).days
    mes_anterior_inicio = inicio_mes - relativedelta(months=1)
    mes_anterior_fim = mes_anterior_inicio + timedelta(days=dias_decorridos_mes)
    receita_mes_atual = _receita_bruta_periodo(inicio_mes, hoje)
    receita_mes_anterior = _receita_bruta_periodo(mes_anterior_inicio, mes_anterior_fim)
    comparacao_mensal = {
        "atual": receita_mes_atual,
        "anterior": receita_mes_anterior,
        "variacao": _variacao_percentual(receita_mes_atual, receita_mes_anterior),
    }

    inicio_ano_atual = hoje.replace(month=1, day=1)
    inicio_ano_anterior = inicio_ano_atual - relativedelta(years=1)
    fim_ano_anterior = inicio_ano_anterior + (hoje - inicio_ano_atual)
    receita_ano_atual = _receita_bruta_periodo(inicio_ano_atual, hoje)
    receita_ano_anterior = _receita_bruta_periodo(inicio_ano_anterior, fim_ano_anterior)
    comparacao_anual = {
        "atual": receita_ano_atual,
        "anterior": receita_ano_anterior,
        "variacao": _variacao_percentual(receita_ano_atual, receita_ano_anterior),
    }

    # --- Gráfico de lucro por mês (últimos 12 meses) -----------------------------
    labels_lucro = []
    valores_lucro = []
    cursor_lucro = inicio_periodo
    for _ in range(12):
        fim_mes_cursor = cursor_lucro + relativedelta(months=1) - timedelta(days=1)
        if fim_mes_cursor > hoje:
            fim_mes_cursor = hoje
        vendas_cursor = Venda.objects.filter(
            status="fechada", fechada_em__date__gte=cursor_lucro, fechada_em__date__lte=fim_mes_cursor
        )
        receita_cursor = _receita_bruta_periodo(cursor_lucro, fim_mes_cursor)
        cmv_cursor = ItemVenda.objects.filter(venda__in=vendas_cursor).aggregate(
            total=Sum(F("quantidade") * F("produto__preco_custo"))
        )["total"] or Decimal("0")
        despesas_cursor = ContaPagar.objects.filter(
            vencimento__gte=cursor_lucro, vencimento__lte=fim_mes_cursor
        ).exclude(categoria__nome="Compra de estoque").aggregate(
            total=Sum("valor"))["total"] or Decimal("0")
        lucro_cursor = receita_cursor - cmv_cursor - despesas_cursor
        labels_lucro.append(f"{MESES_PT[cursor_lucro.month - 1]}/{str(cursor_lucro.year)[2:]}")
        valores_lucro.append(float(lucro_cursor))
        cursor_lucro = cursor_lucro + relativedelta(months=1)

    # --- Receita diária (mês atual, dia a dia) -----------------------------------
    por_dia = (
        Venda.objects.filter(status="fechada", fechada_em__date__gte=inicio_mes, fechada_em__date__lte=hoje)
        .annotate(dia=TruncDay("fechada_em"))
        .values("dia")
        .annotate(valor=Sum(F("itens__quantidade") * F("itens__preco_unitario")))
    )
    mapa_dia = {item["dia"].date(): item["valor"] for item in por_dia if item["dia"]}
    labels_dia = []
    valores_dia = []
    cursor_dia = inicio_mes
    while cursor_dia <= hoje:
        labels_dia.append(cursor_dia.strftime("%d/%m"))
        valores_dia.append(float(mapa_dia.get(cursor_dia, 0)))
        cursor_dia += timedelta(days=1)

    # --- Receita semanal (últimas 9 semanas) -------------------------------------
    inicio_semana_atual = hoje - timedelta(days=hoje.weekday())
    inicio_semanas = inicio_semana_atual - timedelta(weeks=8)
    por_semana = (
        Venda.objects.filter(status="fechada", fechada_em__date__gte=inicio_semanas)
        .annotate(semana=TruncWeek("fechada_em"))
        .values("semana")
        .annotate(valor=Sum(F("itens__quantidade") * F("itens__preco_unitario")))
    )
    mapa_semana = {item["semana"].date(): item["valor"] for item in por_semana if item["semana"]}
    labels_semana = []
    valores_semana = []
    cursor_semana = inicio_semanas
    for _ in range(9):
        labels_semana.append(cursor_semana.strftime("%d/%m"))
        valores_semana.append(float(mapa_semana.get(cursor_semana, 0)))
        cursor_semana += timedelta(weeks=1)

    # --- Fluxo de caixa completo: entradas x saídas (dia/semana/mês/ano) ---------
    def _fluxo_por_bucket(trunc_func, campo, inicio, quantidade, passo):
        """Agrupa MovimentoCaixa (entrada/saída) por um recorte de tempo."""
        dados = (
            MovimentoCaixa.objects.filter(data__gte=inicio)
            .annotate(bucket=trunc_func("data"))
            .values("bucket", "tipo")
            .annotate(total=Sum("valor"))
        )
        mapa = {}
        for item in dados:
            if item["bucket"] is None:
                continue
            chave = item["bucket"] if not hasattr(item["bucket"], "date") else item["bucket"].date()
            mapa.setdefault(chave, {"entrada": 0, "saida": 0})
            mapa[chave][item["tipo"]] = float(item["total"])

        labels, entradas, saidas = [], [], []
        cursor = inicio
        for _ in range(quantidade):
            valores = mapa.get(cursor, {"entrada": 0, "saida": 0})
            labels.append(campo(cursor))
            entradas.append(valores["entrada"])
            saidas.append(valores["saida"])
            cursor = passo(cursor)
        return labels, entradas, saidas

    fluxo_dia_labels, fluxo_dia_entradas, fluxo_dia_saidas = _fluxo_por_bucket(
        TruncDay, lambda d: d.strftime("%d/%m"), inicio_mes,
        (hoje - inicio_mes).days + 1, lambda d: d + timedelta(days=1),
    )
    fluxo_semana_labels, fluxo_semana_entradas, fluxo_semana_saidas = _fluxo_por_bucket(
        TruncWeek, lambda d: d.strftime("%d/%m"), inicio_semanas, 9, lambda d: d + timedelta(weeks=1),
    )
    fluxo_mes_labels, fluxo_mes_entradas, fluxo_mes_saidas = _fluxo_por_bucket(
        TruncMonth, lambda d: f"{MESES_PT[d.month - 1]}/{str(d.year)[2:]}", inicio_periodo, 12,
        lambda d: d + relativedelta(months=1),
    )

    anos_com_dado = sorted(set(
        d.year for d in MovimentoCaixa.objects.values_list("data", flat=True)
    )) or [hoje.year]
    inicio_ano_dados = date(anos_com_dado[0], 1, 1)
    qtd_anos = hoje.year - anos_com_dado[0] + 1
    fluxo_ano_labels, fluxo_ano_entradas, fluxo_ano_saidas = _fluxo_por_bucket(
        TruncYear, lambda d: str(d.year), inicio_ano_dados, qtd_anos,
        lambda d: d + relativedelta(years=1),
    )

    # --- Heatmap de vendas (dia da semana x hora, últimos 12 meses) --------------
    DIAS_SEMANA_PT = ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sáb"]
    heatmap_raw = (
        vendas_periodo
        .annotate(
            dia_semana=ExtractWeekDay("fechada_em", tzinfo=tz_atual),
            hora=ExtractHour("fechada_em", tzinfo=tz_atual),
        )
        .values("dia_semana", "hora")
        .annotate(total=Count("id"))
    )
    matriz_heatmap = {(item["dia_semana"], item["hora"]): item["total"] for item in heatmap_raw}
    heatmap_max = max(matriz_heatmap.values(), default=0)

    heatmap = []
    for dia_idx in range(1, 8):  # convenção do Django: 1=domingo ... 7=sábado
        celulas = []
        for hora in range(24):
            valor = matriz_heatmap.get((dia_idx, hora), 0)
            intensidade = (valor / heatmap_max) if heatmap_max else 0
            alpha = round(0.06 + intensidade * 0.85, 2) if valor else 0.06
            celulas.append({
                "hora": hora,
                "valor": valor,
                "cor": f"rgba(139,108,242,{alpha})",
            })
        heatmap.append({"label": DIAS_SEMANA_PT[dia_idx - 1], "celulas": celulas})

    # --- Central de alertas -------------------------------------------------------
    alertas = []
    PRIORIDADE = {"urgente": 0, "atencao": 1, "info": 2}

    contas_vencidas_pagar = ContaPagar.objects.filter(status="pendente", vencimento__lt=hoje)
    contas_vencidas_receber = ContaReceber.objects.filter(status="pendente", vencimento__lt=hoje)
    total_vencido_pagar = contas_vencidas_pagar.aggregate(total=Sum("valor"))["total"] or Decimal("0")
    total_vencido_receber = contas_vencidas_receber.aggregate(total=Sum("valor"))["total"] or Decimal("0")
    qtd_vencidas = contas_vencidas_pagar.count() + contas_vencidas_receber.count()
    if qtd_vencidas:
        alertas.append({
            "nivel": "urgente",
            "titulo": f"{qtd_vencidas} conta(s) vencida(s)",
            "descricao": f"R$ {total_vencido_pagar} a pagar e R$ {total_vencido_receber} a receber em atraso.",
            "link": "/admin/financeiro/contapagar/",
        })

    produtos_zerados = Produto.objects.filter(ativo=True, estoque_atual=0).count()
    produtos_baixo_nao_zerado = Produto.objects.filter(
        ativo=True, estoque_atual__gt=0, estoque_atual__lte=F("estoque_minimo")
    ).count()
    if produtos_zerados:
        alertas.append({
            "nivel": "urgente",
            "titulo": f"{produtos_zerados} produto(s) zerado(s)",
            "descricao": "Sem nenhuma unidade em estoque no momento.",
            "link": "/admin/catalogo/produto/",
        })
    if produtos_baixo_nao_zerado:
        alertas.append({
            "nivel": "atencao",
            "titulo": f"{produtos_baixo_nao_zerado} produto(s) com estoque baixo",
            "descricao": "Abaixo do estoque mínimo configurado.",
            "link": "/admin/catalogo/produto/",
        })

    if produtos_parados:
        alertas.append({
            "nivel": "atencao",
            "titulo": f"{len(produtos_parados)} produto(s) parado(s)",
            "descricao": f"Sem venda há {config_estoque.dias_produto_parado}+ dias.",
            "link": "#produtos-parados",
        })

    if meta_contexto and not meta_contexto["atingida"]:
        ultimo_dia_mes = (inicio_mes + relativedelta(months=1) - timedelta(days=1)).day
        dias_passados_pct = (hoje.day / ultimo_dia_mes * 100) if ultimo_dia_mes else 0
        if meta_contexto["progresso_percentual"] < dias_passados_pct - 15:
            alertas.append({
                "nivel": "atencao",
                "titulo": "Meta do mês em risco",
                "descricao": (
                    f"Já se passaram {dias_passados_pct:.0f}% do mês, mas só "
                    f"{meta_contexto['progresso_percentual']:.0f}% da meta foi atingida."
                ),
                "link": "#meta-mes",
            })

    ultima_sessao_fechada = CaixaSessao.objects.filter(fechada_em__isnull=False).order_by("-fechada_em").first()
    if ultima_sessao_fechada and ultima_sessao_fechada.diferenca is not None and abs(ultima_sessao_fechada.diferenca) > 20:
        alertas.append({
            "nivel": "urgente" if abs(ultima_sessao_fechada.diferenca) > 50 else "atencao",
            "titulo": "Diferença no último fechamento de caixa",
            "descricao": f"Caixa #{ultima_sessao_fechada.pk} fechou com diferença de R$ {ultima_sessao_fechada.diferenca}.",
            "link": "/admin/financeiro/caixasessao/",
        })

    # --- Pedidos em andamento (vendas abertas, ainda não finalizadas) -----------
    agora = timezone.now()
    vendas_abertas_qs = (
        Venda.objects.filter(status="aberta")
        .select_related("cliente", "vendedor")
        .order_by("aberta_em")
    )
    pedidos_andamento = []
    for v in vendas_abertas_qs:
        tempo_aberto = agora - v.aberta_em
        horas_aberto = tempo_aberto.total_seconds() / 3600
        pedidos_andamento.append({
            "id": v.id,
            "cliente": v.cliente.nome if v.cliente else "Sem cliente identificado",
            "vendedor": v.vendedor.username if v.vendedor else "—",
            "total": v.total,
            "aberta_em": v.aberta_em,
            "horas_aberto": horas_aberto,
        })
    qtd_pedidos_andamento = len(pedidos_andamento)

    if any(p["horas_aberto"] >= 2 for p in pedidos_andamento):
        antigos = sum(1 for p in pedidos_andamento if p["horas_aberto"] >= 2)
        alertas.append({
            "nivel": "atencao",
            "titulo": f"{antigos} pedido(s) em andamento há mais de 2 horas",
            "descricao": "Provavelmente um carrinho esquecido — finalize ou cancele.",
            "link": "/admin/vendas/venda/?status__exact=aberta",
        })

    from core.backup_views import dias_desde_ultimo_backup
    LIMITE_DIAS_BACKUP = 3
    dias_backup = dias_desde_ultimo_backup()
    if dias_backup is None:
        alertas.append({
            "nivel": "atencao",
            "titulo": "Nenhum backup encontrado ainda",
            "descricao": "Baixe um backup agora, e considere agendar o comando automático.",
            "link": "/backup/",
        })
    elif dias_backup >= LIMITE_DIAS_BACKUP:
        alertas.append({
            "nivel": "urgente" if dias_backup >= LIMITE_DIAS_BACKUP * 2 else "atencao",
            "titulo": f"Backup desatualizado — {dias_backup} dia(s) sem backup novo",
            "descricao": "Baixe um backup agora, ou confira se o agendamento automático ainda está rodando.",
            "link": "/backup/",
        })

    alertas.sort(key=lambda a: PRIORIDADE.get(a["nivel"], 9))

    # --- Giro e cobertura de estoque (últimos 90 dias) --------------------------
    JANELA_GIRO_DIAS = 90
    inicio_giro = hoje - timedelta(days=JANELA_GIRO_DIAS)
    vendidos_por_produto = {
        item["produto_id"]: item["total"]
        for item in ItemVenda.objects.filter(
            venda__status="fechada", venda__fechada_em__date__gte=inicio_giro
        ).values("produto_id").annotate(total=Sum("quantidade"))
    }
    giro_cobertura = []
    for p in Produto.objects.filter(ativo=True, estoque_atual__gt=0).order_by("-estoque_atual")[:200]:
        vendido = vendidos_por_produto.get(p.id, 0)
        giro = round(vendido / p.estoque_atual, 2) if p.estoque_atual else None
        media_diaria = vendido / JANELA_GIRO_DIAS
        cobertura_dias = round(p.estoque_atual / media_diaria) if media_diaria > 0 else None
        giro_cobertura.append({
            "produto": p, "vendido_90d": vendido, "giro": giro, "cobertura_dias": cobertura_dias,
        })
    # produtos com giro (venderam algo) primeiro, ordenados por quem gira mais rápido
    giro_cobertura.sort(key=lambda x: (x["giro"] is None, -(x["giro"] or 0)))
    giro_cobertura = giro_cobertura[:10]

    # --- Aniversariantes do mês ---------------------------------------------------
    aniversariantes = sorted(
        Cliente.objects.filter(data_nascimento__month=hoje.month, anonimizado=False),
        key=lambda c: c.data_nascimento.day,
    )

    contexto = {
        "total_valor": total_valor,
        "total_pedidos": total_pedidos,
        "ticket_medio": ticket_medio,
        "produtos_vendidos": produtos_vendidos,
        "total_a_receber": total_a_receber,
        "qtd_a_receber": qtd_a_receber,
        "labels_mes": labels_mes,
        "valores_mes": valores_mes,
        "top_produtos": top_produtos,
        "labels_hora": labels_hora,
        "valores_hora": valores_hora,
        "saldo_atual": saldo_atual,
        "labels_fluxo": labels_fluxo,
        "valores_fluxo": valores_fluxo,
        "proximos_vencimentos": proximos_vencimentos,
        "dre": {
            "receita_bruta": receita_bruta,
            "valor_devolucoes": valor_devolucoes,
            "receita_liquida": receita_liquida,
            "cmv": cmv,
            "margem_bruta": margem_bruta,
            "despesas_operacionais": despesas_operacionais,
            "lucro_liquido": lucro_liquido,
        },
        "indicadores": indicadores,
        "total_produtos_catalogo": total_produtos_catalogo,
        "meta": meta_contexto,
        "produtos_reestoque": produtos_reestoque,
        "comissoes": comissoes,
        "curva_abc": curva_abc,
        "produtos_parados": produtos_parados,
        "dias_produto_parado": config_estoque.dias_produto_parado,
        "comparacao_mensal": comparacao_mensal,
        "comparacao_anual": comparacao_anual,
        "labels_lucro": labels_lucro,
        "valores_lucro": valores_lucro,
        "labels_dia": labels_dia,
        "valores_dia": valores_dia,
        "labels_semana": labels_semana,
        "valores_semana": valores_semana,
        "fluxo_dia_labels": fluxo_dia_labels,
        "fluxo_dia_entradas": fluxo_dia_entradas,
        "fluxo_dia_saidas": fluxo_dia_saidas,
        "fluxo_semana_labels": fluxo_semana_labels,
        "fluxo_semana_entradas": fluxo_semana_entradas,
        "fluxo_semana_saidas": fluxo_semana_saidas,
        "fluxo_mes_labels": fluxo_mes_labels,
        "fluxo_mes_entradas": fluxo_mes_entradas,
        "fluxo_mes_saidas": fluxo_mes_saidas,
        "fluxo_ano_labels": fluxo_ano_labels,
        "fluxo_ano_entradas": fluxo_ano_entradas,
        "fluxo_ano_saidas": fluxo_ano_saidas,
        "heatmap": heatmap,
        "horas_range": list(range(24)),
        "alertas": alertas,
        "pedidos_andamento": pedidos_andamento,
        "qtd_pedidos_andamento": qtd_pedidos_andamento,
        "produtos_menos_vendidos": produtos_menos_vendidos,
        "giro_cobertura": giro_cobertura,
        "aniversariantes": aniversariantes,
        "hoje_dia": hoje.day,
    }
    return contexto
